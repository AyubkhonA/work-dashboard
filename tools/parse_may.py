#!/usr/bin/env python3
"""Real parser for Advanced Ortho Lab monthly statement.
Reads the raw QuickBooks invoice + AR exports, applies per-office sales tax,
and emits a single JSON the dashboard consumes. NO mock data."""

import json
import sys
from openpyxl import load_workbook

DOWNLOADS = "/Users/khan/Downloads"
RAW_PREMIER = f"{DOWNLOADS}/Report_from_Advanced_Orthodontics_Lab_of_San_Francisco (2).xlsx"
RAW_CHILDREN = f"{DOWNLOADS}/Report_from_Advanced_Orthodontics_Lab_of_San_Francisco (1).xlsx"
# AR aging report — the AR total is whatever THIS uploaded file contains (not forced to a reference).
AR_REPORT = f"{DOWNLOADS}/Report_from_Advanced_Orthodontics_Lab_of_San_Francisco (3).xlsx"

# Column indices (1-based) in the RAW invoice exports
C_TYPE = 5    # E  "Invoice"
C_DATE = 7    # G
C_NUM = 9     # I
C_MEMO = 11   # K  (clean item name e.g. "Bands", "Digital Model*")
C_NAME = 13   # M  office
C_PATIENT = 15  # O
C_ITEM = 17   # Q
C_QTY = 19    # S
C_PRICE = 21  # U

# Per-office tax rates (matched by city substring). Default 8.5%.
PREMIER_RATES = {"Vacaville": 8.125, "Yuba City": 7.25}
CHILDREN_RATES = {"Antioch": 9.75, "Glendora": 10.25, "San Diego": 7.25, "Oxnard": 7.75}
DEFAULT_RATE = 8.5


def tax_rate_for(office, company):
    table = PREMIER_RATES if company == "premier" else CHILDREN_RATES
    for city, rate in table.items():
        if city in office:
            return rate
    return DEFAULT_RATE


def num(v):
    return isinstance(v, (int, float))


def parse_invoices(path, company):
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    offices = {}        # office -> {pre_tax, invoices set, lines}
    products = {}       # item -> {orders, units, revenue}
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, C_TYPE).value != "Invoice":
            continue
        office = ws.cell(r, C_NAME).value
        if not office:
            continue
        office = str(office).strip()
        qty = ws.cell(r, C_QTY).value or 0
        price = ws.cell(r, C_PRICE).value
        # mirror =ROUND(IF(ISNUMBER(price), qty*price, qty),5)
        amount = round(qty * price, 5) if num(price) else round(qty, 5)

        o = offices.setdefault(office, {"pre_tax": 0.0, "invoices": set(), "lines": 0})
        o["pre_tax"] += amount
        o["lines"] += 1
        inv = ws.cell(r, C_NUM).value
        if inv is not None:
            o["invoices"].add(inv)

        item = ws.cell(r, C_MEMO).value or ws.cell(r, C_ITEM).value or "Unknown"
        item = str(item).strip()
        patient = ws.cell(r, C_PATIENT).value
        p = products.setdefault(item, {"orders": 0, "units": 0, "revenue": 0.0, "cases": set()})
        p["orders"] += 1
        p["units"] += qty if num(qty) else 0
        p["revenue"] += amount
        if patient:  # a "case" = a distinct patient who got this appliance
            p["cases"].add((office, str(patient).strip()))

    out = []
    for office, d in offices.items():
        rate = tax_rate_for(office, company)
        pre = round(d["pre_tax"], 2)
        with_tax = round(pre + pre * rate / 100.0, 2)
        out.append({
            "office": office,
            "company": company,
            "invoices": len(d["invoices"]),
            "lines": d["lines"],
            "preTax": pre,
            "taxRate": rate,
            "tax": round(with_tax - pre, 2),
            "withTax": with_tax,
        })
    out.sort(key=lambda x: -x["withTax"])
    return out, products


def ar_bucket(days):
    if days is None or not num(days) or days <= 0:
        return "Current"
    if days <= 30:
        return "1-30"
    if days <= 60:
        return "31-60"
    if days <= 90:
        return "61-90"
    return "90+"


MONTHS = ["", "Jan", "Feb", "Mar", "Apr", "May", "Jun",
          "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]


def month_label(key):
    """'2026-05' -> 'May 2026'."""
    try:
        y, m = key.split("-")
        return f"{MONTHS[int(m)]} {y}"
    except Exception:
        return key


def ar_aggregate(pairs):
    """pairs = list of (customer, balance) -> office/group breakdown."""
    by_office, pg, cg = {}, 0.0, 0.0
    for cust, bal in pairs:
        by_office[cust] = by_office.get(cust, 0.0) + bal
        if cust.startswith("Children"):
            cg += bal
        else:
            pg += bal
    total = sum(b for _, b in pairs)
    ranked = sorted(
        [{"office": o, "balance": round(b, 2)} for o, b in by_office.items()],
        key=lambda x: -x["balance"],
    )
    return {
        "total": round(total, 2),
        "premierGroup": round(pg, 2),
        "childrenGroup": round(cg, 2),
        "openInvoices": len(pairs),
        "offices": [dict(x, share=round(x["balance"] / total * 100, 1) if total else 0.0)
                    for x in ranked],
    }


def parse_ar(path):
    """Header-based: detect Customer / Open Balance / Aging / Date columns by name.
    Emits the 'all' aggregate plus a per-invoice-month breakdown so the dashboard can
    show everything in the file, or filter to a single month."""
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    hdr = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if isinstance(v, str):
            hdr[v.strip()] = c
    c_cust = hdr.get("Customer", 2)
    c_open = hdr.get("Open Balance", 14)
    c_age = hdr.get("Aging")   # may be absent
    c_date = hdr.get("Date")   # invoice date -> month bucket

    rows = []  # (cust, bal, days, month_key)
    max_date = None
    for r in range(2, ws.max_row + 1):
        cust = ws.cell(r, c_cust).value
        bal = ws.cell(r, c_open).value
        if not cust or not num(bal):  # skip blank-customer QB total rows
            continue
        days = ws.cell(r, c_age).value if c_age else None
        d = ws.cell(r, c_date).value if c_date else None
        if hasattr(d, "strftime") and (max_date is None or d > max_date):
            max_date = d
        mkey = d.strftime("%Y-%m") if hasattr(d, "strftime") else "undated"
        rows.append((str(cust).strip(), float(bal), days, mkey))

    base = ar_aggregate([(c, b) for c, b, _, _ in rows])
    total = base["total"]

    buckets = {"Current": 0.0, "1-30": 0.0, "31-60": 0.0, "61-90": 0.0, "90+": 0.0}
    for c, b, days, _ in rows:
        buckets[ar_bucket(days)] += b

    monthrows = {}
    for c, b, _, mkey in rows:
        monthrows.setdefault(mkey, []).append((c, b))
    by_month = {m: ar_aggregate(p) for m, p in monthrows.items()}
    months = sorted(
        [{"key": m, "label": month_label(m), "total": by_month[m]["total"],
          "invoices": by_month[m]["openInvoices"]} for m in by_month],
        key=lambda x: x["key"], reverse=True,
    )

    return {
        **base,
        "asOf": max_date.strftime("%Y-%m-%d") if max_date else None,
        "buckets": [{"bucket": k, "balance": round(v, 2),
                     "pct": round(v / total * 100, 1)} for k, v in buckets.items()],
        "months": months,
        "byMonth": by_month,
    }


def main():
    prem_offices, prem_products = parse_invoices(RAW_PREMIER, "premier")
    child_offices, child_products = parse_invoices(RAW_CHILDREN, "children")

    prem_pre = round(sum(o["preTax"] for o in prem_offices), 2)
    prem_tax = round(sum(o["withTax"] for o in prem_offices), 2)
    child_pre = round(sum(o["preTax"] for o in child_offices), 2)
    child_tax = round(sum(o["withTax"] for o in child_offices), 2)

    # merge products (exclude Bands & Digital Model per spec)
    products = {}
    for src in (prem_products, child_products):
        for item, d in src.items():
            p = products.setdefault(item, {"orders": 0, "units": 0, "revenue": 0.0, "cases": set()})
            p["orders"] += d["orders"]
            p["units"] += d["units"]
            p["revenue"] += d["revenue"]
            p["cases"] |= d["cases"]

    def excluded(name):
        n = name.lower()
        return "band" in n or "digital model" in n

    top_products = sorted(
        [{"item": k, "orders": v["orders"], "units": v["units"],
          "cases": len(v["cases"]), "revenue": round(v["revenue"], 2)}
         for k, v in products.items() if not excluded(k)],
        key=lambda x: -x["orders"],
    )[:5]

    ar = parse_ar(AR_REPORT)

    data = {
        "month": "May", "year": 2026,
        "summary": {
            "premierPreTax": prem_pre, "premierWithTax": prem_tax,
            "childrenPreTax": child_pre, "childrenWithTax": child_tax,
            "monthlyTotalWithTax": round(prem_tax + child_tax, 2),
            "totalTax": round((prem_tax - prem_pre) + (child_tax - child_pre), 2),
            "arTotal": ar["total"], "arPremierGroup": ar["premierGroup"],
            "arChildrenGroup": ar["childrenGroup"], "openInvoices": ar["openInvoices"],
        },
        "offices": prem_offices + child_offices,
        "ar": ar,
        "topProducts": top_products,
    }

    out_path = "/Users/khan/work-dashboard/client/src/data/may2026.json"
    import os
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    with open(out_path, "w") as f:
        json.dump(data, f, indent=2)

    # ---- validation report ----
    REF = {"premier": 75021, "children": 3533, "total": 78553, "ar": 148852}
    print("=== PARSED MAY 2026 (REAL FILES) ===")
    print(f"Premier+Sedation  pre-tax ${prem_pre:>12,.2f}   with-tax ${prem_tax:>12,.2f}   (ref ${REF['premier']:,})")
    print(f"Children's Choice pre-tax ${child_pre:>12,.2f}   with-tax ${child_tax:>12,.2f}   (ref ${REF['children']:,})")
    print(f"Monthly total (with tax)  ${prem_tax + child_tax:>12,.2f}                 (ref ${REF['total']:,})")
    print(f"AR total open balance     ${ar['total']:>12,.2f}   (from uploaded AR file, as of {ar['asOf']})")
    print(f"  AR Premier group ${ar['premierGroup']:,.2f}  |  AR Children group ${ar['childrenGroup']:,.2f}")
    print(f"Offices: {len(prem_offices)} Premier/Sedation + {len(child_offices)} Children's = {len(data['offices'])}")
    print(f"Top products: {', '.join(p['item'][:24] for p in top_products)}")

    def check(label, got, ref, tol=0.02):
        diff = abs(got - ref) / ref
        flag = "OK" if diff <= tol else f"OFF by {diff*100:.1f}%"
        print(f"  [{flag}] {label}: got {got:,.0f} vs ref {ref:,}")

    print("--- reconciliation (2% tolerance) ---")
    check("Premier with tax", prem_tax, REF["premier"])
    check("Children with tax", child_tax, REF["children"])
    check("Monthly total", prem_tax + child_tax, REF["total"])
    print(f"\nWrote {out_path}")


if __name__ == "__main__":
    main()
