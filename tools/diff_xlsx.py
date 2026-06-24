#!/usr/bin/env python3
"""Attribute-level diff between a generated xlsx and the golden template.
Compares, per cell: value/formula, number_format, font(name,size,bold,italic,color),
alignment(horizontal,vertical), border(sides+style), fill. Plus column widths,
row heights, freeze panes, sheet view. Reports up to LIMIT diffs per category."""
import sys
import openpyxl

LIMIT = 40

def colcolor(c):
    col = c.font.color
    if col is None: return None
    try:
        rgb = col.rgb
        if isinstance(rgb, str): return ('rgb', rgb)
    except Exception:
        pass
    # theme color
    return ('theme', getattr(col, 'theme', None), getattr(col, 'tint', None))

def borders(c):
    out = {}
    for s in ('left','right','top','bottom'):
        b = getattr(c.border, s)
        if b is not None and b.style:
            out[s] = b.style
    return out

def norm_h(h):
    # horizontal 'general' is the DEFAULT and renders identically to unset (None):
    # text left, numbers right. ExcelJS can't emit 'general', so treat them equal.
    return None if h in (None, 'general') else h

def cellsig(c):
    return {
        'v': c.value,
        'fmt': c.number_format,
        'font': (c.font.name, float(c.font.size) if c.font.size else None, bool(c.font.bold), bool(c.font.italic), colcolor(c)),
        'al': (norm_h(c.alignment.horizontal), c.alignment.vertical),
        'bd': tuple(sorted(borders(c).items())),
        'fill': c.fill.patternType,
    }

def main(gen_fn, gold_fn):
    g = openpyxl.load_workbook(gen_fn)
    G = openpyxl.load_workbook(gold_fn)
    gs, Gs = g.active, G.active
    diffs = {'dims':[], 'freeze':[], 'colw':[], 'rowh':[], 'cell':[]}

    if (gs.max_row, gs.max_column) != (Gs.max_row, Gs.max_column):
        diffs['dims'].append(f"gen={gs.max_row}x{gs.max_column} gold={Gs.max_row}x{Gs.max_column}")
    if gs.freeze_panes != Gs.freeze_panes:
        diffs['freeze'].append(f"gen={gs.freeze_panes!r} gold={Gs.freeze_panes!r}")

    cols = [openpyxl.utils.get_column_letter(i) for i in range(1, max(gs.max_column, Gs.max_column)+1)]
    for c in cols:
        gw = gs.column_dimensions[c].width if c in gs.column_dimensions else None
        Gw = Gs.column_dimensions[c].width if c in Gs.column_dimensions else None
        if (gw or 0) != (Gw or 0) and abs((gw or 0)-(Gw or 0)) > 0.001:
            diffs['colw'].append(f"col {c}: gen={gw} gold={Gw}")

    maxr = max(gs.max_row, Gs.max_row)
    for r in range(1, maxr+1):
        gh = gs.row_dimensions[r].height if r in gs.row_dimensions else None
        Gh = Gs.row_dimensions[r].height if r in Gs.row_dimensions else None
        if (gh or 0) != (Gh or 0) and abs((gh or 0)-(Gh or 0)) > 0.001:
            if len(diffs['rowh']) < LIMIT:
                diffs['rowh'].append(f"row {r}: gen_h={gh} gold_h={Gh}")

    for r in range(1, maxr+1):
        for c in cols:
            gc = gs[f"{c}{r}"]; Gc = Gs[f"{c}{r}"]
            a, b = cellsig(gc), cellsig(Gc)
            if a != b:
                d = {k: (a[k], b[k]) for k in a if a[k] != b[k]}
                if len(diffs['cell']) < LIMIT:
                    diffs['cell'].append((f"{c}{r}", d))

    total = sum(len(v) for v in diffs.values())
    print(f"==== DIFF {gen_fn}  vs  {gold_fn} ====")
    print(f"dims:{len(diffs['dims'])} freeze:{len(diffs['freeze'])} colw:{len(diffs['colw'])} rowh:{len(diffs['rowh'])} cell:(showing<= {LIMIT})")
    for k in ('dims','freeze','colw'):
        for x in diffs[k]: print(f"  [{k}] {x}")
    for x in diffs['rowh'][:12]: print(f"  [rowh] {x}")
    # summarize cell diffs by which attributes differ
    from collections import Counter
    attr_counter = Counter()
    # recount ALL cell diffs (not just stored) for accurate totals
    allcell = 0
    for r in range(1, maxr+1):
        for c in cols:
            a, b = cellsig(gs[f"{c}{r}"]), cellsig(Gs[f"{c}{r}"])
            if a != b:
                allcell += 1
                for k in a:
                    if a[k] != b[k]: attr_counter[k]+=1
    print(f"  [cell] TOTAL cell diffs: {allcell}  by-attr: {dict(attr_counter)}")
    for coord, d in diffs['cell'][:LIMIT]:
        print(f"   {coord}: " + " | ".join(f"{k} gen={v[0]!r} gold={v[1]!r}" for k,v in d.items()))
    return allcell

if __name__ == '__main__':
    main(sys.argv[1], sys.argv[2])
