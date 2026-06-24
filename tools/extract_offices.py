#!/usr/bin/env python3
"""Extract the office/line-item structure from a golden statement xlsx into JSON
that matches what buildCompanyWorkbook(offices, ...) expects. Used to drive the
format-fidelity diff: generate from this data, diff vs the same golden -> only
FORMAT differences surface (data is identical)."""
import sys, json, re
import openpyxl

def extract(fn):
    wb = openpyxl.load_workbook(fn)  # formulas
    ws = wb.active
    offices = []
    cur = None
    r = 2
    while r <= ws.max_row:
        A = ws[f"A{r}"].value
        B = ws[f"B{r}"].value
        if A == 'TOTAL':
            break
        if B and isinstance(B, str) and B.startswith('Total '):
            # subtotal row; next row is tax -> read rate
            tax = ws[f"S{r+1}"].value
            rate = None
            if isinstance(tax, str):
                m = re.search(r'\*\s*([0-9.]+)\s*%', tax)
                if m: rate = float(m.group(1))
            if cur is not None:
                cur['taxRate'] = rate
            cur = None
            r += 2  # skip subtotal + tax
            continue
        if B and isinstance(B, str):
            # banner -> new office
            cur = {'office': B, 'taxRate': None, 'items': []}
            offices.append(cur)
            r += 1
            continue
        # item row
        E = ws[f"E{r}"].value
        G = ws[f"G{r}"].value
        K = ws[f"K{r}"].value
        M = ws[f"M{r}"].value
        O = ws[f"O{r}"].value
        Q = ws[f"Q{r}"].value
        date = None
        if E is not None and hasattr(E, 'year'):
            date = f"{E.year:04d}-{E.month:02d}-{E.day:02d}"
        elif E is not None:
            date = str(E)
        if cur is not None and (E is not None or G is not None or M is not None or O is not None):
            cur['items'].append({
                'date': date,
                'num': None if G is None else str(G),
                'patient': None if K is None else str(K),
                'item': None if M is None else str(M),
                'qty': O,
                'price': Q,
            })
        r += 1
    return offices

if __name__ == '__main__':
    fn = sys.argv[1]
    offices = extract(fn)
    n_items = sum(len(o['items']) for o in offices)
    print(json.dumps(offices), file=sys.stdout)
    print(f"# {fn}: {len(offices)} offices, {n_items} items", file=sys.stderr)
    for o in offices:
        print(f"#   {o['office']!r} rate={o['taxRate']} items={len(o['items'])}", file=sys.stderr)
